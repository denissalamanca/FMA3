#!/usr/bin/env python3
"""reconcile_demo.py — the weekly demo-forward reconciliation harness (DEMO_FORWARD_PLAN §6D).

Ingests a FableBookNative run's own outputs — the per-hour telemetry CSV
(FMA3_native_hourly.csv) and the Strategy-Tester deals report (.xlsx) — and
produces the weekly checkpoint: plumbing health, worst-mark drawdown path +
crisis windows, friction decomposition, and (when a matched record reference is
supplied) growth-factor retention + native record->tick k.

Design notes (honest scope):
  * Self-contained metrics (plumbing / DD / friction) need ONLY the run's own
    output — they work for the LIVE demo period.
  * The 'vs record' block (retention, k) needs a matched record curve. Past the
    frozen horizon (2025-12-31) there is no golden to compare against
    (WARMSTART_DESIGN §7.4), so --record is optional; without it those rows are
    skipped, not faked.
  * min-ML is read from the telemetry IF the margin/ML field has been added
    (DEMO_FORWARD_PLAN §6C.1); until then it reports 'not logged'.

Usage:
  reconcile_demo.py --telemetry FMA3_native_hourly.csv --report ReportTester-*.xlsx \
                    [--record hrisk1_s160_curve.parquet] [--window 2024-07-29:2024-08-16 ...] \
                    [--json out.json]

Validated 2026-07-16 against report _43 (real-tick IC 2023-2025): net €134,862,
worst-mark DD 18.76%, swap -€26.5k, crisis DDs Aug-24 -15.5% / Apr-25 -13.1%.
"""
from __future__ import annotations
import argparse, json, sys
from pathlib import Path
import numpy as np
import pandas as pd


def load_telemetry(path: str):
    """Returns (hourly H rows, per-symbol P rows). P rows exist only from the
    2026-07 EA build onward; older telemetry has none, so callers must cope."""
    df = pd.read_csv(path)
    h = df[df["rec"] == "H"].copy()
    h["dt"] = pd.to_datetime(h["ts"], unit="s")
    p = df[df["rec"] == "P"].copy() if "rec" in df else df.iloc[0:0]
    if len(p):
        p["dt"] = pd.to_datetime(p["ts"], unit="s")
    return (h.sort_values("dt").reset_index(drop=True),
            p.sort_values("dt").reset_index(drop=True) if len(p) else p)


def position_fidelity(p: pd.DataFrame, band: float) -> dict:
    """DEMO_GO_NOGO #2 — the honest fidelity measure.

    NOT "held == want": the executor has a rebalance dead-band (InpRebalBand,
    BookExec.mqh:~276) and deliberately does NOT retrade while
    ||want|-|held||/|held| <= band. Scoring held==want would read ~0% BY DESIGN.

    What we actually assert is the executor's own invariant, per leg:
      * want == 0        -> held must be flat
      * sign(want) != sign(held) -> violation (never intended after a pass)
      * else             -> drift must be within the band
    Legs legitimately deferred (market closed / unsized: no quote) are excluded
    from the invariant but counted, because a leg stuck deferred IS a failure.
    """
    if not len(p):
        return {"status": "no P rows — telemetry predates the 2026-07 build"}
    q = p.copy()
    for c in ("want", "held", "defer"):
        q[c] = pd.to_numeric(q[c], errors="coerce")
    q["deferred"] = q["defer"].fillna(0).astype(int) > 0
    act = q[~q["deferred"]].copy()

    aw, ah = act["want"].abs(), act["held"].abs()
    flat_ok = (aw < 1e-9) & (ah < 1e-9)
    sign_bad = (aw >= 1e-9) & (ah >= 1e-9) & (np.sign(act["want"]) != np.sign(act["held"]))
    # drift only where we hold something and want something
    both = (aw >= 1e-9) & (ah >= 1e-9)
    drift = pd.Series(0.0, index=act.index)
    drift[both] = (aw[both] - ah[both]).abs() / ah[both]
    # want!=0 but flat (or vice-versa) is a violation unless deferred
    mismatch_flat = ((aw >= 1e-9) & (ah < 1e-9)) | ((aw < 1e-9) & (ah >= 1e-9))
    act["ok"] = flat_ok | (both & ~sign_bad & (drift <= band + 1e-12))
    act.loc[mismatch_flat | sign_bad, "ok"] = False

    per_bar = act.groupby("dt")["ok"].all()
    stuck = q[q["deferred"]].groupby("sym").size().sort_values(ascending=False)
    return {
        "band": band,
        "n_leg_bars": int(len(act)),
        "leg_fidelity": float(act["ok"].mean()) if len(act) else float("nan"),
        "bar_fidelity": float(per_bar.mean()) if len(per_bar) else float("nan"),
        "n_bars": int(len(per_bar)),
        "max_drift": float(drift.max()) if len(drift) else 0.0,
        "n_sign_violations": int(sign_bad.sum()),
        "worst_legs": act[~act["ok"]].groupby("sym").size().sort_values(
            ascending=False).head(5).to_dict(),
        "deferred_leg_bars": int(q["deferred"].sum()),
        "most_deferred": stuck.head(3).to_dict(),
    }


def ftmo_envelope(h: pd.DataFrame, initial: float) -> dict:
    """DEMO_GO_NOGO #2 — the 5%-daily / 10%-max-loss rule envelope, computed off
    the day_anchor the EA now logs (Guardian's own anchor, so the harness and
    the breaker agree on what a 'day' is instead of the harness guessing)."""
    if "day_anchor" not in h or pd.to_numeric(h["day_anchor"], errors="coerce").isna().all():
        return {"status": "no day_anchor — telemetry predates the 2026-07 build"}
    q = h.copy()
    q["day_anchor"] = pd.to_numeric(q["day_anchor"], errors="coerce")
    eq = pd.to_numeric(q.get("worst_eq", q["equity"]), errors="coerce").fillna(q["equity"])
    anch = q["day_anchor"].replace(0.0, np.nan)
    max_loss = (initial - eq) / initial                 # fraction below the initial balance
    # Guardian maintains g_fedAnchor ONLY when the breaker is armed (Guardian.mqh:80 —
    # InpDailyStopX<=0 returns early, "OFF: no state"). So on an IC run (dailyStopX=0)
    # day_anchor is 0 on every row and the daily-5% half is genuinely UN-MEASURABLE —
    # not zero, not passing. The max-loss half needs only the initial balance, so it
    # still stands. (Found by running against real run-44 telemetry; the synthetic
    # fixture had a live anchor on every row and could not have caught this.)
    if anch.isna().all():
        return {
            "initial": initial,
            "status": "daily anchor never set — breaker DISARMED (InpDailyStopX=0). "
                      "Daily-5% un-measurable here; it needs the FTMO preset.",
            "worst_daily_loss": None,
            "worst_total_loss": float(max_loss.max()),
            "breaches_10pct_max_loss": int((max_loss > 0.10).sum()),
            "PASS": None,
        }
    daily_loss = (anch - eq) / anch                     # fraction below the day's anchor
    return {
        "initial": initial,
        "worst_daily_loss": float(daily_loss.max()) if daily_loss.notna().any() else None,
        "n_hours_past_5pct_daily": int((daily_loss > 0.05).sum()),
        "n_days_breaching_5pct": int(q.loc[daily_loss > 0.05, "dt"].dt.date.nunique()),
        "worst_total_loss": float(max_loss.max()),
        "breaches_10pct_max_loss": int((max_loss > 0.10).sum()),
        "PASS": bool((max_loss > 0.10).sum() == 0),
    }


def load_deals(path: str) -> pd.DataFrame:
    import openpyxl  # lazy: only needed when a report is passed
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = [list(r) for r in wb.active.iter_rows(values_only=True)]
    hi = next((i for i, r in enumerate(rows)
               if r and "Deal" in [str(x) for x in r]
               and "Swap" in [str(x) for x in r] and "Balance" in [str(x) for x in r]), None)
    if hi is None:
        raise ValueError("deals table header (Deal/Swap/Balance) not found in report")
    cols = [str(x) for x in rows[hi]]
    body = []
    for r in rows[hi + 1:]:
        if r[0] is None:
            break
        body.append(r)
    d = pd.DataFrame(body, columns=cols[:len(body[0])])
    d = d[pd.to_numeric(d["Deal"], errors="coerce").notna()].copy()
    for c in ("Commission", "Swap", "Profit", "Balance"):
        if c in d:
            d[c] = pd.to_numeric(d[c], errors="coerce")
    d["dt"] = pd.to_datetime(d["Time"], errors="coerce")
    return d


def worst_mark_dd(eq: pd.Series) -> float:
    """Worst peak-to-trough on an hourly-equity series (mark-to-market proxy)."""
    if len(eq) < 2:
        return float("nan")
    peak = np.maximum.accumulate(eq.values)
    return float(((eq.values - peak) / peak).min())


def window_slice(h: pd.DataFrame, a: str, b: str) -> pd.DataFrame:
    return h[(h["dt"] >= pd.Timestamp(a)) & (h["dt"] <= pd.Timestamp(b))]


def reconcile(telemetry: str, report: str | None, record: str | None,
              windows: list[str], band: float = 0.25,
              initial: float | None = None) -> dict:
    h, p = load_telemetry(telemetry)
    out: dict = {"span": [str(h["dt"].iloc[0]), str(h["dt"].iloc[-1])],
                 "n_hours": int(len(h))}

    # --- plumbing health (self-contained) ---
    # NOTE (go/no-go 2026-07-16): sc_mm is the SC-sleeve SIGNAL self-check, NOT
    # position fidelity. 'fires' is the CoreTrigger SEGMENT count, NOT the FTMO
    # breaker. Both kept under their true names; the real breaker count is
    # n_stops (g_fedNStops), now in the hourly row rather than deinit-only.
    def _last(col):
        return (int(pd.to_numeric(h[col], errors="coerce").dropna().iloc[-1])
                if col in h and pd.to_numeric(h[col], errors="coerce").notna().any() else None)

    warm_col = pd.to_numeric(h["warm"], errors="coerce") if "warm" in h else None
    out["plumbing"] = {
        "sc_selfcheck_max": int(h["sc_mm"].max()) if "sc_mm" in h else None,   # signal self-check, NOT fidelity
        "core_segments_max": int(h["fires"].max()) if "fires" in h else None,  # CoreTrigger segments, NOT breaker
        "unready_max": int(h["unready"].max()) if "unready" in h else None,
        "trading_from": (str(h[h["trading"] == 1]["dt"].min())
                         if "trading" in h and (h["trading"] == 1).any() else "NEVER"),
        "breaker_stops": _last("n_stops"),   # the REAL breaker count (g_fedNStops)
        # #3: `trading` alone LIES — it is 1 on a silent cold start too. warm==0
        # while trading==1 means the book is computing from now, not resumed.
        "warm": (None if warm_col is None or warm_col.isna().all()
                 else ("cold for the WHOLE span — silent cold start"
                       if (warm_col == 0).all() else
                       "warm" if (warm_col == 1).all() else
                       f"cold until {h.loc[warm_col.eq(1).idxmax(), 'dt']}")),
    }
    out["position_fidelity"] = position_fidelity(p, band)
    if initial is not None:
        out["ftmo_envelope"] = ftmo_envelope(h, initial)

    # --- equity + drawdown (self-contained) ---
    # The §5 kill line (28%) is a WORST-MARK drawdown. Hourly equity samples the
    # hour boundary and understates it, so prefer the EA's own FED_WorstMarkEquity
    # when the telemetry carries it; fall back to hourly equity (and say so).
    eq = h["equity"]
    weq = pd.to_numeric(h["worst_eq"], errors="coerce") if "worst_eq" in h else None
    has_worst = weq is not None and weq.notna().any() and (weq > 0).any()
    out["equity"] = {
        "start": float(eq.iloc[0]), "end": float(eq.iloc[-1]),
        "return": float(eq.iloc[-1] / eq.iloc[0] - 1),
        "worst_mark_dd_hourly": worst_mark_dd(eq),
        "dd_basis": ("worst-mark (FED_WorstMarkEquity)" if has_worst
                     else "hourly equity — UNDERSTATES the 28% line (telemetry predates the 2026-07 build)"),
    }
    if has_worst:
        peak = np.maximum.accumulate(eq.values)
        out["equity"]["worst_mark_dd"] = float(((weq.values - peak) / peak).min())
    # min-ML if the margin field was added (§6C.1); over rows with positions held
    # (margin_level == 0 means no open positions — the cold/flat state, not a real ML)
    ml_col = next((c for c in h.columns if c.lower() in ("margin_level", "ml", "marginlevel")), None)
    if ml_col:
        active = pd.to_numeric(h[ml_col], errors="coerce")
        active = active[active > 0]
        out["equity"]["min_ML"] = (float(active.min()) if len(active) else "no positions held")
    else:
        out["equity"]["min_ML"] = "not logged (add margin/ML to telemetry — §6C.1)"

    # --- crisis / weekly windows ---
    out["windows"] = {}
    for w in windows:
        a, b = w.split(":")
        ws = window_slice(h, a, b)
        if len(ws) < 2:
            out["windows"][w] = {"n_hours": int(len(ws)), "note": "insufficient data"}
            continue
        we = ws["equity"]
        out["windows"][w] = {
            "n_hours": int(len(ws)),
            "return": float(we.iloc[-1] / we.iloc[0] - 1),
            "worst_mark_dd": worst_mark_dd(we),
        }

    # --- friction (needs the deals report) ---
    if report:
        d = load_deals(report)
        swap, comm = float(d["Swap"].sum()), float(d["Commission"].sum())
        prof = float(d["Profit"].sum())
        out["friction"] = {
            "n_deals": int(len(d)),
            "swap": swap, "commission": comm,
            "gross_profit_col_sum": prof,
            "net_incl_swap_comm": prof + swap + comm,
            "swap_pct_of_gross": (swap / prof * 100 if prof else None),
        }

    # --- vs record (optional; skipped past the frozen horizon) ---
    if record:
        rc = pd.read_parquet(record)
        eqcol = "equity" if "equity" in rc else rc.columns[0]
        rl = rc[eqcol].resample("1h").last().dropna()
        rl = rl[(rl.index >= h["dt"].iloc[0]) & (rl.index <= h["dt"].iloc[-1])]
        if len(rl) > 1:
            run_gf = eq.iloc[-1] / eq.iloc[0]
            rec_gf = rl.iloc[-1] / rl.iloc[0]
            out["vs_record"] = {
                "run_growth_factor": float(run_gf),
                "record_growth_factor": float(rec_gf),
                "growth_retention": float(run_gf / rec_gf),
            }
            # native k per window (run worst-mark DD / record worst-mark DD)
            wc = "worst" if "worst" in rc else eqcol
            rw = rc[wc].resample("1h").min().dropna()
            out["vs_record"]["k_by_window"] = {}
            for w in windows:
                a, b = w.split(":")
                run_dd = out["windows"].get(w, {}).get("worst_mark_dd")
                rws = rw[(rw.index >= pd.Timestamp(a)) & (rw.index <= pd.Timestamp(b))]
                rls = rl[(rl.index >= pd.Timestamp(a)) & (rl.index <= pd.Timestamp(b))]
                if run_dd and len(rls) > 1:
                    peak = rls.cummax()
                    rec_dd = float(((rws.reindex(rls.index) - peak) / peak).min())
                    out["vs_record"]["k_by_window"][w] = {
                        "run_dd": run_dd, "record_dd": rec_dd,
                        "k": (run_dd / rec_dd if rec_dd else None)}
        else:
            out["vs_record"] = "record does not overlap the run window (past the frozen horizon — expected for the live demo)"
    else:
        out["vs_record"] = "no --record supplied (skipped, not faked)"

    return out


def _fmt(out: dict) -> str:
    L = []
    L.append(f"DEMO RECONCILIATION — span {out['span'][0]} → {out['span'][1]} ({out['n_hours']:,} hrs)")
    p = out["plumbing"]
    L.append(f"  PLUMBING: sc-selfcheck max={p['sc_selfcheck_max']} · core-segments={p['core_segments_max']} · trading from {p['trading_from']}")
    L.append(f"    breaker stops (g_fedNStops)={p['breaker_stops']} · warm: {p['warm']}")
    f = out["position_fidelity"]
    if "status" in f:
        L.append(f"  FIDELITY: {f['status']}")
    else:
        L.append(f"  FIDELITY: bar {f['bar_fidelity']:.2%} · leg {f['leg_fidelity']:.3%} "
                 f"(invariant: within band {f['band']}, sign-correct; NOT held==want) "
                 f"· max drift {f['max_drift']:.2f} · sign-violations {f['n_sign_violations']}")
        if f["worst_legs"]:
            L.append(f"    worst legs: {f['worst_legs']}")
        if f["deferred_leg_bars"]:
            L.append(f"    deferred leg-bars: {f['deferred_leg_bars']} · most: {f['most_deferred']}")
    e = out["equity"]
    dd = e.get("worst_mark_dd", e["worst_mark_dd_hourly"])
    L.append(f"  EQUITY:   {e['start']:,.0f} → {e['end']:,.0f}  ret {e['return']:+.1%}  DD {dd:.2%}  min ML {e['min_ML']}")
    L.append(f"    DD basis: {e['dd_basis']}")
    if "ftmo_envelope" in out:
        v = out["ftmo_envelope"]
        if "status" in v:
            L.append(f"  FTMO:     {v['status']}")
            if v.get("worst_total_loss") is not None:
                L.append(f"            (max-loss half still computable: worst total "
                         f"{v['worst_total_loss']:.2%} · 10% breaches {v['breaches_10pct_max_loss']})")
        else:
            wd = v["worst_daily_loss"]
            wd_s = f"{wd:.2%}" if wd is not None else "n/a"
            L.append(f"  FTMO:     worst daily {wd_s} (5% breach days: {v['n_days_breaching_5pct']}) · "
                     f"worst total {v['worst_total_loss']:.2%} · 10% breaches {v['breaches_10pct_max_loss']} · "
                     f"{'PASS' if v['PASS'] else 'FAIL'}")
    if "friction" in out:
        f = out["friction"]
        L.append(f"  FRICTION: {f['n_deals']:,} deals · swap {f['swap']:,.0f} · comm {f['commission']:,.0f} · net {f['net_incl_swap_comm']:,.0f}")
    for w, v in out["windows"].items():
        if "worst_mark_dd" in v:
            L.append(f"  WINDOW {w}: ret {v['return']:+.1%}  worst-mark DD {v['worst_mark_dd']:.2%}")
    if isinstance(out.get("vs_record"), dict):
        vr = out["vs_record"]
        L.append(f"  vs RECORD: growth retention {vr.get('growth_retention', float('nan')):.2f}×")
        for w, k in vr.get("k_by_window", {}).items():
            L.append(f"     k[{w}] = {k['k']:.2f}× (run {k['run_dd']:.2%} / record {k['record_dd']:.2%})")
    else:
        L.append(f"  vs RECORD: {out['vs_record']}")
    return "\n".join(L)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Demo-forward weekly reconciliation harness")
    ap.add_argument("--telemetry", required=True, help="FMA3_native_hourly.csv")
    ap.add_argument("--report", help="ReportTester-*.xlsx (for friction)")
    ap.add_argument("--record", help="matched record curve parquet (optional; omit past the frozen horizon)")
    ap.add_argument("--window", action="append", default=[], help="A:B window(s), e.g. 2024-07-29:2024-08-16")
    ap.add_argument("--band", type=float, default=0.25,
                    help="InpRebalBand — the executor's churn dead-band. Fidelity is measured "
                         "against THIS, not against held==want (see position_fidelity).")
    ap.add_argument("--initial", type=float,
                    help="initial balance -> enables the FTMO 5%%/10%% envelope (e.g. 100000)")
    ap.add_argument("--json", help="also write the full result as JSON")
    a = ap.parse_args(argv)
    out = reconcile(a.telemetry, a.report, a.record, a.window, a.band, a.initial)
    print(_fmt(out))
    if a.json:
        Path(a.json).write_text(json.dumps(out, indent=1, default=str))
        print(f"\n[full JSON -> {a.json}]")
    return 0


if __name__ == "__main__":
    sys.exit(main())
